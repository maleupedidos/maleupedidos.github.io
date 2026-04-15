"""
Extrae TODAS las ventas del Excel MALEU 2026 y genera ventas.json
Formato unificado para la PWA Admin.
Ejecutar: python extract_ventas.py
"""
import openpyxl, json, sys
from datetime import datetime

import shutil, tempfile
_SRC = 'C:/Users/tadeu/OneDrive - Económicas - UBA/Maleu 2026/Administración/MALEU 2026.xlsm'
PATH = tempfile.mktemp(suffix='.xlsm')
shutil.copy2(_SRC, PATH)
OUT = 'C:/Users/tadeu/maleupedidos.github.io/data/ventas.json'

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True, keep_vba=False)
ventas = []

def safe(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    return str(v).strip()

def num(v):
    try: return float(v) if v else 0
    except: return 0

def fmt_date(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if v: return str(v).strip()
    return ''

# ═══ VENTAS HOME / PILAR / CAPITAL FEDERAL (misma estructura) ═══
for canal, sheet_name in [('Home', 'Ventas Home'), ('Pilar', 'Ventas Pilar'), ('Capital Federal', 'Ventas Capital Federal')]:
    ws = wb[sheet_name]
    if ws.max_row <= 1: continue
    for r in range(2, ws.max_row + 1):
        cliente = safe(ws.cell(r, 8).value)
        if not cliente: continue
        # Usar fecha de ENTREGA (col 48) en vez de pedido (col 4)
        MESES_VALIDOS = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
        fecha_entrega = ws.cell(r, 48).value
        fecha = fmt_date(fecha_entrega) if fecha_entrega else fmt_date(ws.cell(r, 4).value)
        mes_raw = safe(ws.cell(r, 49).value)
        mes_entrega = mes_raw if mes_raw in MESES_VALIDOS else safe(ws.cell(r, 5).value)
        sem_raw = num(ws.cell(r, 50).value)
        sem_entrega = sem_raw if sem_raw > 0 else num(ws.cell(r, 6).value)
        facturado = num(ws.cell(r, 20).value) or num(ws.cell(r, 14).value)
        if facturado == 0: continue

        ventas.append({
            'canal': canal,
            'fecha': fecha,
            'mes': mes_entrega,
            'sem': int(sem_entrega),
            'cliente': cliente,
            'estado': safe(ws.cell(r, 11).value),
            'fp': safe(ws.cell(r, 12).value),
            'ep': safe(ws.cell(r, 13).value),
            'facturado': facturado,
            'efectivo': num(ws.cell(r, 16).value),
            'transferencia': num(ws.cell(r, 17).value),
            'costo': num(ws.cell(r, 40).value),
            'margen': num(ws.cell(r, 41).value),
        })

# ═══ VENTAS CLUBES (estructura diferente) ═══
ws = wb['Ventas Clubes']
if ws.max_row > 1:
    for r in range(2, ws.max_row + 1):
        cliente = safe(ws.cell(r, 8).value)
        if not cliente: continue
        club = safe(ws.cell(r, 9).value)
        facturado = num(ws.cell(r, 23).value) or num(ws.cell(r, 17).value)
        if facturado == 0: continue

        ventas.append({
            'canal': 'Clubes',
            'fecha': fmt_date(ws.cell(r, 4).value),
            'mes': safe(ws.cell(r, 5).value),
            'sem': int(num(ws.cell(r, 6).value)),
            'cliente': cliente + (' (' + club + ')' if club else ''),
            'estado': safe(ws.cell(r, 14).value),
            'fp': safe(ws.cell(r, 15).value),
            'ep': safe(ws.cell(r, 16).value),
            'facturado': facturado,
            'efectivo': num(ws.cell(r, 19).value),
            'transferencia': num(ws.cell(r, 20).value),
            'costo': num(ws.cell(r, 32).value),
            'margen': num(ws.cell(r, 33).value),
        })

# ═══ VENTAS RED (headers en row 3, data from row 4) ═══
ws = wb['Ventas Red']
if ws.max_row > 3:
    for r in range(4, ws.max_row + 1):
        cliente = safe(ws.cell(r, 6).value)
        if not cliente: continue
        facturado = num(ws.cell(r, 13).value)
        if facturado == 0: continue

        ventas.append({
            'canal': 'Red',
            'fecha': fmt_date(ws.cell(r, 2).value),
            'mes': safe(ws.cell(r, 3).value),
            'sem': int(num(ws.cell(r, 4).value)),
            'cliente': cliente,
            'estado': 'Entregado' if safe(ws.cell(r, 11).value) in ('Si', 'SI', '1') else 'Pendiente',
            'fp': 'Transferencia' if num(ws.cell(r, 15).value) > 0 else 'Efectivo',
            'ep': 'Cobrado',
            'facturado': facturado,
            'efectivo': num(ws.cell(r, 14).value),
            'transferencia': num(ws.cell(r, 15).value),
            'costo': 0,
            'margen': 0,
        })

# ═══ VENTAS B2B (headers en row 3, data from row 4) ═══
ws = wb['Ventas B2B']
if ws.max_row > 3:
    for r in range(4, ws.max_row + 1):
        cliente = safe(ws.cell(r, 6).value)
        if not cliente: continue
        facturado = num(ws.cell(r, 9).value)
        if facturado == 0: continue

        ventas.append({
            'canal': 'B2B',
            'fecha': fmt_date(ws.cell(r, 2).value),
            'mes': safe(ws.cell(r, 3).value),
            'sem': int(num(ws.cell(r, 4).value)),
            'cliente': cliente,
            'estado': 'Entregado' if safe(ws.cell(r, 7).value) in ('Si', 'SI', '1') else 'Pendiente',
            'fp': 'Transferencia' if num(ws.cell(r, 11).value) > 0 else 'Efectivo',
            'ep': 'Cobrado',
            'facturado': facturado,
            'efectivo': num(ws.cell(r, 10).value),
            'transferencia': num(ws.cell(r, 11).value),
            'costo': 0,
            'margen': 0,
        })

# ═══ VENTAS CATERING (headers en row 3, data from row 4) ═══
ws = wb['Ventas Catering']
if ws.max_row > 3:
    for r in range(4, ws.max_row + 1):
        cliente = safe(ws.cell(r, 6).value)
        if not cliente: continue
        facturado = num(ws.cell(r, 9).value)
        if facturado == 0: continue

        ventas.append({
            'canal': 'Catering',
            'fecha': fmt_date(ws.cell(r, 2).value),
            'mes': safe(ws.cell(r, 3).value),
            'sem': int(num(ws.cell(r, 4).value)),
            'cliente': cliente,
            'estado': 'Entregado',
            'fp': 'Transferencia' if num(ws.cell(r, 12).value) > 0 else 'Efectivo',
            'ep': 'Cobrado',
            'facturado': facturado,
            'efectivo': num(ws.cell(r, 11).value),
            'transferencia': num(ws.cell(r, 12).value),
            'costo': num(ws.cell(r, 18).value),
            'margen': num(ws.cell(r, 23).value),
        })

# Sort by date desc
ventas.sort(key=lambda x: x['fecha'], reverse=True)

# Write JSON
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(ventas, f, ensure_ascii=False)

print(f'OK: {len(ventas)} ventas exportadas a {OUT}')

# Summary
canales = {}
meses = {}
for v in ventas:
    canales[v['canal']] = canales.get(v['canal'], 0) + 1
    meses[v['mes']] = meses.get(v['mes'], 0) + 1
print('\nPor canal:', json.dumps(canales, ensure_ascii=False))
print('Por mes:', json.dumps(meses, ensure_ascii=False))
total = sum(v['facturado'] for v in ventas)
print(f'Facturación total: ${total:,.0f}')
